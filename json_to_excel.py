import os
import json
import logging
from typing import Dict, Any, List, Optional, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def convert_json_to_excel(json_file_path: str, excel_output_path: Optional[str] = None) -> str:
    logger.info(f"Converting JSON file {json_file_path} to Excel")

    if not excel_output_path:
        base_name = os.path.splitext(os.path.basename(json_file_path))[0]
        excel_output_path = f"{base_name}.xlsx"

    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            doc_data = json.load(f)
    except Exception as e:
        logger.error(f"Error loading JSON file: {e}")
        raise

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    for page in doc_data.get("pages", []):
        page_number = page.get("page_number", 0)
        page_content = page.get("content", {})

        sheet_name = f"Page {page_number}"
        ws = wb.create_sheet(title=sheet_name)

        format_page_worksheet(ws, page_content, page_number)

    try:
        wb.save(excel_output_path)
        logger.info(f"Excel file saved to {excel_output_path}")
        return excel_output_path
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        raise

def format_page_worksheet(ws: Worksheet, page_content: Dict[str, Any], page_number: int) -> None:
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    document_type = page_content.get("document_type", "Unknown Document Type")
    page_metadata = page_content.get("page_metadata", {})

    ws.cell(row=1, column=1, value=f"Document Type: {document_type}")
    ws.cell(row=1, column=1).font = title_font

    if "header" in page_metadata and page_metadata["header"]:
        ws.cell(row=2, column=1, value=f"Header: {page_metadata['header']}")

    ws.cell(row=3, column=1, value=f"Page: {page_number}")

    current_row = 5

    tables = page_content.get("tables", [])
    if tables:
        for table_idx, table in enumerate(tables):
            table_title = table.get("table_title", f"Table {table_idx+1}")
            headers = table.get("headers", [])
            data = table.get("data", [])

            ws.cell(row=current_row, column=1, value=table_title)
            ws.cell(row=current_row, column=1).font = title_font
            current_row += 1

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(len(str(header)) + 2, 12)
            
            current_row += 1

            for row_data in data:
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.border = border
                
                current_row += 1

            current_row += 2

    sections = page_content.get("sections", [])
    if sections:
        for section_idx, section in enumerate(sections):
            section_type = section.get("section_type", "")
            section_title = section.get("section_title", f"Section {section_idx+1}")
            content = section.get("content", "")

            if section_type == "table":
                continue

            ws.cell(row=current_row, column=1, value=section_title)
            ws.cell(row=current_row, column=1).font = title_font
            current_row += 1

            if section_type == "text":
                ws.cell(row=current_row, column=1, value=content)
                current_row += 2
            elif section_type == "form":
                if isinstance(content, dict):
                    for key, value in content.items():
                        ws.cell(row=current_row, column=1, value=key)
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        ws.cell(row=current_row, column=2, value=value)
                        current_row += 1
                    current_row += 1
                elif isinstance(content, list):
                    for item in content:
                        if isinstance(item, dict):
                            for key, value in item.items():
                                ws.cell(row=current_row, column=1, value=key)
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=2, value=value)
                                current_row += 1
                    current_row += 1
            elif section_type == "chart":
                ws.cell(row=current_row, column=1, value=f"Chart: {content}")
                current_row += 2

    key_value_pairs = page_content.get("key_value_pairs", {})
    if key_value_pairs:
        ws.cell(row=current_row, column=1, value="Additional Information")
        ws.cell(row=current_row, column=1).font = title_font
        current_row += 1
        
        for key, value in key_value_pairs.items():
            ws.cell(row=current_row, column=1, value=key)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

    for col in range(1, 10):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

def create_pandas_dataframes(page_content: Dict[str, Any]) -> List[Tuple[str, pd.DataFrame]]:
    dataframes = []
    
    tables = page_content.get("tables", [])
    for table_idx, table in enumerate(tables):
        table_title = table.get("table_title", f"Table {table_idx+1}")
        headers = table.get("headers", [])
        data = table.get("data", [])
        
        if headers and data:
            df = pd.DataFrame(data, columns=headers)
            dataframes.append((table_title, df))

    sections = page_content.get("sections", [])
    for section_idx, section in enumerate(sections):
        if section.get("section_type") == "form":
            section_title = section.get("section_title", f"Form {section_idx+1}")
            content = section.get("content", {})
            
            if isinstance(content, dict):
                df = pd.DataFrame([content])
                dataframes.append((section_title, df))
            elif isinstance(content, list) and all(isinstance(x, dict) for x in content):
                df = pd.DataFrame(content)
                dataframes.append((section_title, df))

    key_value_pairs = page_content.get("key_value_pairs", {})
    if key_value_pairs:
        df = pd.DataFrame(key_value_pairs.items(), columns=["Key", "Value"])
        dataframes.append(("Key-Value Pairs", df))
    
    return dataframes

def main(json_file_path: str, excel_output_path: Optional[str] = None) -> str:
    logger.info(f"Starting conversion of {json_file_path} to Excel")
    
    if not excel_output_path:
        base_name = os.path.splitext(os.path.basename(json_file_path))[0]
        excel_output_path = f"{base_name}.xlsx"
    
    try:
        excel_path = convert_json_to_excel(json_file_path, excel_output_path)
        logger.info(f"Successfully converted JSON to Excel: {excel_path}")
        return excel_path
    
    except Exception as e:
        logger.error(f"Error converting JSON to Excel: {e}")
        raise

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Convert extracted JSON data to Excel format")
    parser.add_argument("json_file", help="Path to the JSON file containing extracted data")
    parser.add_argument("--output", help="Path to save the Excel file")
    
    args = parser.parse_args()
    
    main(args.json_file, args.output)