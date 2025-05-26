import cv2
import pandas as pd
import os
import json
import numpy as np
import re
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def detect_signature_patterns(text_blocks, img_shape):
    signature_indicators = []
    
    # Text-based signature indicators
    signature_keywords = [
        'signature', 'signed', 'sign', 'name', 'initial', 'initials',
        'authorized', 'approved', 'verified', 'confirm', 'acknowledge',
        'witness', 'attest', 'certify', 'endorse'
    ]
    
    # Pattern matching for signature-like text
    signature_patterns = [
        r'\b[A-Z][a-z]+ [A-Z][a-z]+\b',  # Full names (e.g., "John Smith")
        r'\b[A-Z]\.[A-Z]\. [A-Z][a-z]+\b',  # Initials + surname (e.g., "J.S. Smith")
        r'\b[A-Z]{2,4}\b',  # Initials only (e.g., "JS", "ABC")
        r'\/s\/ .+',  # Electronic signature format
        r'_+\s*$',  # Underscores (signature lines)
        r'-{3,}',  # Dashes (signature lines)
    ]
    
    for block in text_blocks:
        text = block['text'].strip()
        is_signature = False
        confidence_score = 0
        
        # Check for signature keywords
        if any(keyword in text.lower() for keyword in signature_keywords):
            is_signature = True
            confidence_score += 0.4
        
        # Check for signature patterns
        for pattern in signature_patterns:
            if re.search(pattern, text):
                is_signature = True
                confidence_score += 0.3
        
        # Check for visual characteristics that might indicate signatures
        # Low confidence text (handwritten text often has lower OCR confidence)
        if block['confidence'] < 0.7:
            confidence_score += 0.2
        
        # Check for unusual aspect ratios (signatures often have different proportions)
        aspect_ratio = block['width'] / block['height'] if block['height'] > 0 else 0
        if aspect_ratio > 3 or aspect_ratio < 0.5:
            confidence_score += 0.1
        
        # Check for isolated text blocks (signatures are often standalone)
        if len(text.split()) <= 3 and len(text) > 2:
            confidence_score += 0.2
        
        # Special patterns for common signature formats
        if re.match(r'^[A-Z][a-z]*\.?\s*[A-Z][a-z]*\.?$', text):  # Name patterns
            confidence_score += 0.3
        
        if confidence_score >= 0.5:
            signature_indicators.append({
                'block': block,
                'confidence': confidence_score,
                'type': 'text_signature'
            })
    
    return signature_indicators

def analyze_signature_regions(img, text_blocks):
    signature_regions = []
    
    # Convert to grayscale for analysis
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Look for regions with high variance (potentially handwritten content)
    for block in text_blocks:
        x1, y1 = int(block['min_x']), int(block['min_y'])
        x2, y2 = int(block['max_x']), int(block['max_y'])
        
        # Ensure coordinates are within image bounds
        x1 = max(0, x1 - 10)
        y1 = max(0, y1 - 10)
        x2 = min(gray.shape[1], x2 + 10)
        y2 = min(gray.shape[0], y2 + 10)
        
        if x2 > x1 and y2 > y1:
            region = gray[y1:y2, x1:x2]
            
            if region.size > 0:
                # Calculate variance (signatures often have more irregular patterns)
                variance = np.var(region)
                
                # Calculate edge density (signatures often have more edges)
                edges = cv2.Canny(region, 50, 150)
                edge_density = np.sum(edges > 0) / edges.size
                
                # Combine metrics for signature likelihood
                signature_score = (variance / 1000) + (edge_density * 2)
                
                if signature_score > 0.3:  # Threshold for signature detection
                    signature_regions.append({
                        'block': block,
                        'signature_score': signature_score,
                        'type': 'visual_signature'
                    })
    
    return signature_regions

def improved_table_extraction(image_path, output_dir='./output'):
    os.makedirs(output_dir, exist_ok=True)

    ocr = PaddleOCR(
        use_angle_cls=True, 
        lang='en',
        use_gpu=True,
        show_log=False,
        draw_img_save_dir=None,
        vis_font_path=None
    )

    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f"Could not load image from {image_path}")
    
    try:
        result = ocr.ocr(img, cls=True)
        
        if not result or not result[0]:
            print("No text detected in the image")
            return None

        text_info = []
        for line in result[0]:
            box = line[0]
            text = line[1][0].strip()
            confidence = line[1][1]

            if not text:
                continue

            x_coords = [p[0] for p in box]
            y_coords = [p[1] for p in box]
            
            min_x = min(x_coords)
            max_x = max(x_coords)
            min_y = min(y_coords)
            max_y = max(y_coords)
            
            width = max_x - min_x
            height = max_y - min_y
            
            center_x = (min_x + max_x) / 2
            center_y = (min_y + max_y) / 2
            
            text_info.append({
                'text': text,
                'confidence': confidence,
                'center_x': center_x,
                'center_y': center_y,
                'min_x': min_x,
                'max_x': max_x,
                'min_y': min_y,
                'max_y': max_y,
                'width': width,
                'height': height,
                'box': box
            })

        if not text_info:
            print("No valid text blocks detected")
            return None

        # Detect signatures before processing table structure
        print("Detecting signatures...")
        text_signatures = detect_signature_patterns(text_info, img.shape)
        visual_signatures = analyze_signature_regions(img, text_info)
        
        all_signatures = text_signatures + visual_signatures
        signature_blocks = {block['block']['text']: block for block in all_signatures}
        
        print(f"Found {len(all_signatures)} potential signatures")
        for sig in all_signatures:
            print(f"  - '{sig['block']['text']}' (confidence: {sig.get('confidence', sig.get('signature_score', 0)):.2f})")

        text_info.sort(key=lambda x: x['center_y'])

        rows = []
        current_row = [text_info[0]]
        
        heights = [item['height'] for item in text_info]
        avg_height = sum(heights) / len(heights)
        row_height_threshold = avg_height * 0.8
        
        for i in range(1, len(text_info)):
            current_block = text_info[i]
            reference_block = current_row[0]
            
            y_overlap = min(current_block['max_y'], reference_block['max_y']) - max(current_block['min_y'], reference_block['min_y'])
            
            if y_overlap > 0 or abs(current_block['center_y'] - reference_block['center_y']) < row_height_threshold:
                current_row.append(current_block)
            else:
                current_row.sort(key=lambda x: x['center_x'])
                rows.append(current_row)
                current_row = [current_block]

        if current_row:
            current_row.sort(key=lambda x: x['center_x'])
            rows.append(current_row)

        all_centers_x = [block['center_x'] for row in rows for block in row]
        
        column_centers = []
        if len(all_centers_x) > 10:
            sorted_x = sorted(all_centers_x)
            gaps = [(sorted_x[i+1] - sorted_x[i], i) for i in range(len(sorted_x)-1)]
            gaps.sort(reverse=True)
            
            num_columns = min(7, len(gaps) // 3 + 2)
            
            separators = sorted([sorted_x[gap[1]] for gap in gaps[:num_columns-1]])
            
            column_centers = [sorted_x[0] / 2]  # Start
            for i in range(len(separators)):
                mid_point = (separators[i] + (separators[i+1] if i+1 < len(separators) else sorted_x[-1])) / 2
                column_centers.append(mid_point)
        
        if not column_centers or len(column_centers) < 3:
            num_columns = max(len(row) for row in rows)
            
            min_x = min(block['min_x'] for row in rows for block in row)
            max_x = max(block['max_x'] for row in rows for block in row)
            
            column_width = (max_x - min_x) / num_columns
            column_centers = [min_x + column_width * (i + 0.5) for i in range(num_columns)]

        table_matrix = []
        signature_row_indices = []
        
        for row_idx, row in enumerate(rows):
            row_data = [''] * len(column_centers)
            row_has_signature = False
            
            for block in row:
                distances = [abs(block['center_x'] - center) for center in column_centers]
                closest_col = distances.index(min(distances))
                
                if row_data[closest_col]:
                    row_data[closest_col] += ' ' + block['text']
                else:
                    row_data[closest_col] = block['text']
                
                # Check if this block contains a signature
                if block['text'] in signature_blocks:
                    row_has_signature = True
            
            table_matrix.append(row_data)
            if row_has_signature:
                signature_row_indices.append(row_idx)

        
        headers = []
        header_row_idx = 0
        
        if table_matrix and any(cell for cell in table_matrix[0] if cell.lower() in ['s. no', 's.no', 'sno', 'sl. no', 'serial no']):
            headers = table_matrix[0]
        else:
            non_empty_cols = sum(1 for cell in table_matrix[0] if cell.strip())
            if non_empty_cols >= min(3, len(table_matrix[0])):
                headers = table_matrix[0]
            else:
                headers = [f"Column {i+1}" for i in range(len(column_centers))]
                header_row_idx = -1  # No header row to skip
        
        headers = [h.strip() or f"Column {i+1}" for i, h in enumerate(headers)]
        
        if header_row_idx >= 0:
            data_rows = table_matrix[header_row_idx+1:]
            # Adjust signature row indices to account for removed header
            signature_row_indices = [idx - (header_row_idx + 1) for idx in signature_row_indices if idx > header_row_idx]
        else:
            data_rows = table_matrix
            
        max_cols = len(headers)
        normalized_rows = []
        final_signature_indices = []
        
        for row_idx, row in enumerate(data_rows):
            if not any(cell.strip() for cell in row):
                continue
                
            if len(row) < max_cols:
                row = row + [''] * (max_cols - len(row))
            elif len(row) > max_cols:
                row = row[:max_cols]
                
            normalized_rows.append(row)
            
            # Track signature rows in the final normalized data
            if row_idx in signature_row_indices:
                final_signature_indices.append(len(normalized_rows) - 1)

        if not normalized_rows:
            print("No valid table data extracted")
            return None

        df = pd.DataFrame(normalized_rows, columns=headers)
        
        cleaned_df = df.copy()
        cleaned_signature_indices = final_signature_indices.copy()
        
        i = 0
        while i < len(cleaned_df) - 1:
            current_row = cleaned_df.iloc[i]
            next_row = cleaned_df.iloc[i+1]
            
            current_non_empty = current_row.astype(str).str.strip().ne('').sum()
            next_non_empty = next_row.astype(str).str.strip().ne('').sum()
            
            if next_non_empty <= max_cols // 2 and current_non_empty > next_non_empty:
                for col in cleaned_df.columns:
                    if pd.notna(next_row[col]) and next_row[col].strip():
                        if pd.notna(current_row[col]) and current_row[col].strip():
                            cleaned_df.at[i, col] = f"{current_row[col]} {next_row[col]}"
                        else:
                            cleaned_df.at[i, col] = next_row[col]
    
                if i+1 in cleaned_signature_indices:
                    cleaned_signature_indices.remove(i+1)
                    if i not in cleaned_signature_indices:
                        cleaned_signature_indices.append(i)
                
                cleaned_df = cleaned_df.drop(i+1)
                cleaned_df = cleaned_df.reset_index(drop=True)

                cleaned_signature_indices = [idx if idx <= i else idx - 1 for idx in cleaned_signature_indices]
            else:
                i += 1
        
        if 'S. No' in cleaned_df.columns or any(col for col in cleaned_df.columns if 'no' in col.lower()):
            sno_col = next((col for col in cleaned_df.columns if 'no' in col.lower()), cleaned_df.columns[0])
            
            cleaned_df[sno_col] = pd.to_numeric(cleaned_df[sno_col], errors='coerce')
            
            mask = cleaned_df[sno_col].isna()
            if mask.any():
                valid_sns = cleaned_df.loc[~mask, sno_col].dropna()
                if not valid_sns.empty:
                    last_sn = valid_sns.iloc[-1]
                    counter = last_sn + 1
                    
                    for idx in cleaned_df.index[mask]:
                        cleaned_df.at[idx, sno_col] = counter
                        counter += 1

        csv_path = os.path.join(output_dir, "extracted_table.csv")
        cleaned_df.to_csv(csv_path, index=False)
        print(f"Table saved to CSV: {csv_path}")

        excel_path = os.path.join(output_dir, "extracted_table.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Table"

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for r in dataframe_to_rows(cleaned_df, index=False, header=True):
            ws.append(r)

        for sig_idx in cleaned_signature_indices:
            excel_row = sig_idx + 2  # +1 for header, +1 for 1-indexed
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).fill = green_fill
        
        wb.save(excel_path)
        print(f"Table saved to Excel with signature highlighting: {excel_path}")
        print(f"Highlighted {len(cleaned_signature_indices)} rows with detected signatures")

        txt_path = os.path.join(output_dir, "extracted_table.txt")
        with open(txt_path, 'w', encoding='utf-8') as txt_file:
            txt_file.write("\t".join(cleaned_df.columns) + "\n")
            txt_file.write("-" * 80 + "\n")
            
            for idx, (_, row) in enumerate(cleaned_df.iterrows()):
                prefix = "[SIGNATURE] " if idx in cleaned_signature_indices else ""
                txt_file.write(prefix + "\t".join(str(cell) for cell in row) + "\n")
        
        print(f"Table saved to TXT: {txt_path}")

        json_path = os.path.join(output_dir, "extracted_table.json")
        json_data = {
            "table_data": {
                "headers": cleaned_df.columns.tolist(),
                "rows": cleaned_df.values.tolist(),
                "signature_rows": cleaned_signature_indices
            },
            "signature_analysis": {
                "detected_signatures": [{
                    "text": sig['block']['text'],
                    "confidence": sig.get('confidence', sig.get('signature_score', 0)),
                    "type": sig['type'],
                    "position": {
                        "center_x": float(sig['block']['center_x']),
                        "center_y": float(sig['block']['center_y'])
                    }
                } for sig in all_signatures]
            },
            "raw_data": {
                "text_blocks": [{
                    "text": item["text"],
                    "confidence": float(item["confidence"]),
                    "position": {
                        "center_x": float(item["center_x"]),
                        "center_y": float(item["center_y"]),
                        "box": [[float(p[0]), float(p[1])] for p in item["box"]]
                    },
                    "is_signature": item["text"] in signature_blocks
                } for item in text_info]
            }
        }
        
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)
        
        print(f"Table saved to JSON: {json_path}")

        print("\nExtracted table content:")
        print(cleaned_df)
        
        if cleaned_signature_indices:
            print(f"\nRows with detected signatures (highlighted in green in Excel): {cleaned_signature_indices}")
        
        return cleaned_df, json_data
    
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    image_path = "/home/talgotram/Repos/ioclOCR/input/images/page_3.jpg"
    output_dir = "/home/talgotram/Repos/ioclOCR/output/table_output_3"
    
    try:
        improved_table_extraction(image_path, output_dir)
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()