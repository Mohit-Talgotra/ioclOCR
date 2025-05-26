import cv2
import pandas as pd
import os
import json
import numpy as np
import re
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import time
from typing import List, Dict, Any, Optional

class HuggingFaceTableProcessor:
    """
    HuggingFace-based processor for intelligent table structure detection and reformatting
    """
    
    def __init__(self, api_token: Optional[str] = None, model: str = "microsoft/DialoGPT-large"):
        """
        Initialize the HuggingFace processor
        
        Args:
            api_token: HuggingFace API token (optional for free tier)
            model: HuggingFace model to use for analysis
        """
        self.api_token = api_token
        # Using a model that's good for structured text analysis
        self.model = "microsoft/DialoGPT-large"  # Good for conversational analysis
        # Alternative models you can try:
        # "google/flan-t5-large"  # Good for instruction following
        # "bigscience/bloom-560m"  # Good for text generation and analysis
        # "microsoft/DialoGPT-medium"  # Smaller, faster version
        
        self.headers = {
            "Authorization": f"Bearer {api_token}" if api_token else None,
            "Content-Type": "application/json"
        }
        
        # Remove None headers
        self.headers = {k: v for k, v in self.headers.items() if v is not None}
        
        self.base_url = "https://api-inference.huggingface.co/models"
        
    def _make_api_call(self, payload: dict, max_retries: int = 3) -> dict:
        """
        Make API call to HuggingFace with retry logic
        """
        url = f"{self.base_url}/{self.model}"
        
        for attempt in range(max_retries):
            try:
                response = requests.post(url, headers=self.headers, json=payload, timeout=30)
                
                if response.status_code == 503:
                    # Model is loading, wait and retry
                    wait_time = 20 * (attempt + 1)
                    print(f"Model loading, waiting {wait_time} seconds...")
                    time.sleep(wait_time)
                    continue
                    
                if response.status_code == 200:
                    return response.json()
                else:
                    print(f"API call failed with status {response.status_code}: {response.text}")
                    
            except requests.exceptions.RequestException as e:
                print(f"Request failed (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(5)
                    
        return None
    
    def analyze_table_structure(self, raw_text_blocks: List[Dict]) -> Dict[str, Any]:
        """
        Use HuggingFace model to analyze raw text blocks and determine proper table structure
        """
        # Prepare simplified text data for analysis
        text_content = []
        for i, block in enumerate(raw_text_blocks[:50]):  # Limit to first 50 blocks to avoid token limits
            text_content.append(f"{i}: '{block['text']}' (y:{int(block['center_y'])}, x:{int(block['center_x'])})")
        
        # Create a more structured prompt for better analysis
        analysis_prompt = f"""Analyze these OCR text blocks from a document table:

            {chr(10).join(text_content[:30])}  

            Task: Identify table structure
            1. Which blocks are headers?
            2. What are logical column names?
            3. Which blocks form data rows?
            4. Any special rows (totals, signatures)?

            Response format:
            HEADERS: [list header block numbers]
            COLUMNS: [suggest column names]
            SPECIAL: [list special row block numbers and types]
            STRUCTURE: [brief analysis]"""

        try:
            # Try text generation approach
            payload = {
                "inputs": analysis_prompt,
                "parameters": {
                    "max_new_tokens": 200,
                    "temperature": 0.1,
                    "return_full_text": False
                }
            }
            
            result = self._make_api_call(payload)
            
            if result:
                return self._parse_hf_response(result, raw_text_blocks)
            else:
                print("HuggingFace analysis failed, using fallback")
                return self._fallback_analysis(raw_text_blocks)
                
        except Exception as e:
            print(f"HuggingFace analysis failed: {e}")
            return self._fallback_analysis(raw_text_blocks)
    
    def _parse_hf_response(self, hf_result: dict, raw_text_blocks: List[Dict]) -> Dict[str, Any]:
        """
        Parse HuggingFace response and extract structured information
        """
        try:
            # Handle different response formats
            if isinstance(hf_result, list) and len(hf_result) > 0:
                text_response = hf_result[0].get('generated_text', '')
            elif isinstance(hf_result, dict):
                text_response = hf_result.get('generated_text', str(hf_result))
            else:
                text_response = str(hf_result)
            
            print(f"HF Response: {text_response[:500]}...")  # Debug output
            
            # Parse the response for structured information
            headers = []
            header_blocks = []
            special_blocks = []
            column_names = []
            
            # Extract header block numbers
            if "HEADERS:" in text_response:
                header_section = text_response.split("HEADERS:")[1].split("\n")[0]
                header_nums = re.findall(r'\d+', header_section)
                header_blocks = [int(num) for num in header_nums if int(num) < len(raw_text_blocks)]
            
            # Extract column names
            if "COLUMNS:" in text_response:
                column_section = text_response.split("COLUMNS:")[1].split("\n")[0]
                # Simple extraction of potential column names
                column_names = [name.strip() for name in re.findall(r'[A-Za-z][A-Za-z\s]*', column_section)[:6]]
            
            # Extract special blocks
            if "SPECIAL:" in text_response:
                special_section = text_response.split("SPECIAL:")[1].split("\n")[0]
                special_nums = re.findall(r'\d+', special_section)
                special_blocks = [int(num) for num in special_nums if int(num) < len(raw_text_blocks)]
            
            # If we got header blocks, use their text as column names
            if header_blocks:
                headers = [raw_text_blocks[i]['text'] for i in header_blocks if i < len(raw_text_blocks)]
            elif column_names:
                headers = column_names[:6]  # Limit to reasonable number
            else:
                # Fallback: use first row as headers
                y_positions = sorted(set(block['center_y'] for block in raw_text_blocks))
                if y_positions:
                    first_row_y = y_positions[0]
                    headers = [block['text'] for block in raw_text_blocks 
                             if abs(block['center_y'] - first_row_y) < 20][:6]
            
            # Create data rows structure
            data_rows = []
            processed_blocks = set(header_blocks)
            
            # Group remaining blocks into rows by Y coordinate
            remaining_blocks = [block for i, block in enumerate(raw_text_blocks) 
                              if i not in processed_blocks]
            
            if remaining_blocks:
                remaining_blocks.sort(key=lambda x: x['center_y'])
                
                current_row_blocks = []
                current_y = remaining_blocks[0]['center_y']
                row_threshold = 30  # pixels
                
                for i, block in enumerate(remaining_blocks):
                    if abs(block['center_y'] - current_y) < row_threshold:
                        current_row_blocks.append(i)
                    else:
                        if current_row_blocks:
                            special_type = "signature" if any(j in special_blocks for j in current_row_blocks) else None
                            data_rows.append({
                                "row_number": len(data_rows) + 1,
                                "blocks": current_row_blocks,
                                "special_type": special_type
                            })
                        current_row_blocks = [i]
                        current_y = block['center_y']
                
                # Add last row
                if current_row_blocks:
                    special_type = "signature" if any(j in special_blocks for j in current_row_blocks) else None
                    data_rows.append({
                        "row_number": len(data_rows) + 1,
                        "blocks": current_row_blocks,
                        "special_type": special_type
                    })
            
            return {
                "headers": headers if headers else [f"Column_{i+1}" for i in range(6)],
                "header_blocks": header_blocks,
                "data_rows": data_rows,
                "column_mapping": {},
                "analysis": f"HuggingFace analysis completed. Found {len(headers)} headers, {len(data_rows)} data rows.",
                "raw_response": text_response[:500]  # Store for debugging
            }
            
        except Exception as e:
            print(f"Error parsing HuggingFace response: {e}")
            return self._fallback_analysis(raw_text_blocks)
    
    def _fallback_analysis(self, raw_text_blocks: List[Dict]) -> Dict[str, Any]:
        """
        Enhanced fallback analysis when HuggingFace is not available
        """
        # Improved heuristic-based analysis
        headers = []
        header_blocks = []
        
        # Sort blocks by Y position
        sorted_blocks = sorted(enumerate(raw_text_blocks), key=lambda x: x[1]['center_y'])
        
        # Try to identify headers (usually in the first few rows)
        y_positions = sorted(set(block['center_y'] for block in raw_text_blocks))
        first_row_y = y_positions[0] if y_positions else 0
        
        # Look for common header patterns
        header_keywords = ['name', 'date', 'amount', 'description', 'type', 'status', 'id', 'number', 'total']
        
        for i, block in enumerate(raw_text_blocks):
            # Check if in first row
            if abs(block['center_y'] - first_row_y) < 25:
                # Check if text looks like a header
                text_lower = block['text'].lower().strip()
                is_header = (
                    any(keyword in text_lower for keyword in header_keywords) or
                    len(text_lower.split()) <= 3 or  # Short text
                    text_lower.replace('.', '').replace(' ', '').isalpha()  # Mostly alphabetic
                )
                
                if is_header:
                    headers.append(block['text'])
                    header_blocks.append(i)
        
        # If no headers found, create generic ones
        if not headers:
            # Estimate number of columns from first few rows
            first_rows = []
            for y_pos in y_positions[:3]:
                row_blocks = [block for block in raw_text_blocks 
                             if abs(block['center_y'] - y_pos) < 25]
                first_rows.append(len(row_blocks))
            
            num_cols = max(first_rows) if first_rows else 5
            headers = [f"Column_{i+1}" for i in range(min(num_cols, 8))]
        
        return {
            "headers": headers,
            "header_blocks": header_blocks,
            "data_rows": [],
            "column_mapping": {},
            "analysis": "Enhanced fallback analysis - HuggingFace not available"
        }

def detect_signature_patterns(text_blocks, img_shape):
    """
    Detect potential signature patterns in text blocks and image regions
    """
    signature_indicators = []
    
    # Text-based signature indicators
    signature_keywords = [
        'signature', 'signed', 'sign', 'name', 'initial', 'initials',
        'authorized', 'approved', 'verified', 'confirm', 'acknowledge',
        'witness', 'attest', 'certify', 'endorse'
    ]
    
    # Pattern matching for signature-like text
    signature_patterns = [
        r'\b[A-Z][a-z]+ [A-Z][a-z]+\b',  # Full names (e.g., "John Smith")
        r'\b[A-Z]\.[A-Z]\. [A-Z][a-z]+\b',  # Initials + surname (e.g., "J.S. Smith")
        r'\b[A-Z]{2,4}\b',  # Initials only (e.g., "JS", "ABC")
        r'\/s\/ .+',  # Electronic signature format
        r'_+\s*$',  # Underscores (signature lines)
        r'-{3,}',  # Dashes (signature lines)
    ]
    
    for block in text_blocks:
        text = block['text'].strip()
        is_signature = False
        confidence_score = 0
        
        # Check for signature keywords
        if any(keyword in text.lower() for keyword in signature_keywords):
            is_signature = True
            confidence_score += 0.4
        
        # Check for signature patterns
        for pattern in signature_patterns:
            if re.search(pattern, text):
                is_signature = True
                confidence_score += 0.3
        
        # Check for visual characteristics that might indicate signatures
        if block['confidence'] < 0.7:
            confidence_score += 0.2
        
        # Check for unusual aspect ratios
        aspect_ratio = block['width'] / block['height'] if block['height'] > 0 else 0
        if aspect_ratio > 3 or aspect_ratio < 0.5:
            confidence_score += 0.1
        
        # Check for isolated text blocks
        if len(text.split()) <= 3 and len(text) > 2:
            confidence_score += 0.2
        
        # Special patterns for common signature formats
        if re.match(r'^[A-Z][a-z]*\.?\s*[A-Z][a-z]*\.?$', text):
            confidence_score += 0.3
        
        if confidence_score >= 0.5:
            signature_indicators.append({
                'block': block,
                'confidence': confidence_score,
                'type': 'text_signature'
            })
    
    return signature_indicators

def analyze_signature_regions(img, text_blocks):
    """
    Analyze image regions for visual signature characteristics
    """
    signature_regions = []
    
    # Convert to grayscale for analysis
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Look for regions with high variance (potentially handwritten content)
    for block in text_blocks:
        x1, y1 = int(block['min_x']), int(block['min_y'])
        x2, y2 = int(block['max_x']), int(block['max_y'])
        
        # Ensure coordinates are within image bounds
        x1 = max(0, x1 - 10)
        y1 = max(0, y1 - 10)
        x2 = min(gray.shape[1], x2 + 10)
        y2 = min(gray.shape[0], y2 + 10)
        
        if x2 > x1 and y2 > y1:
            region = gray[y1:y2, x1:x2]
            
            if region.size > 0:
                # Calculate variance
                variance = np.var(region)
                
                # Calculate edge density
                edges = cv2.Canny(region, 50, 150)
                edge_density = np.sum(edges > 0) / edges.size
                
                # Combine metrics for signature likelihood
                signature_score = (variance / 1000) + (edge_density * 2)
                
                if signature_score > 0.3:
                    signature_regions.append({
                        'block': block,
                        'signature_score': signature_score,
                        'type': 'visual_signature'
                    })
    
    return signature_regions

def enhanced_table_extraction(image_path, output_dir='./output', use_ai=True, hf_api_token=None):
    """
    Enhanced table extraction with HuggingFace-based structure analysis
    """
    os.makedirs(output_dir, exist_ok=True)

    # Initialize HuggingFace processor if requested
    hf_processor = None
    if use_ai:
        try:
            hf_processor = HuggingFaceTableProcessor(api_token=hf_api_token)
            print("HuggingFace processor initialized successfully")
        except Exception as e:
            print(f"Failed to initialize HuggingFace processor: {e}")
            print("Continuing with traditional extraction...")

    ocr = PaddleOCR(
        use_angle_cls=True, 
        lang='en',
        use_gpu=True,
        show_log=False,
        draw_img_save_dir=None,
        vis_font_path=None
    )

    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f"Could not load image from {image_path}")
    
    try:
        result = ocr.ocr(img, cls=True)
        
        if not result or not result[0]:
            print("No text detected in the image")
            return None

        # Extract text information
        text_info = []
        for line in result[0]:
            box = line[0]
            text = line[1][0].strip()
            confidence = line[1][1]

            if not text:
                continue

            x_coords = [p[0] for p in box]
            y_coords = [p[1] for p in box]
            
            min_x = min(x_coords)
            max_x = max(x_coords)
            min_y = min(y_coords)
            max_y = max(y_coords)
            
            width = max_x - min_x
            height = max_y - min_y
            
            center_x = (min_x + max_x) / 2
            center_y = (min_y + max_y) / 2
            
            text_info.append({
                'text': text,
                'confidence': confidence,
                'center_x': center_x,
                'center_y': center_y,
                'min_x': min_x,
                'max_x': max_x,
                'min_y': min_y,
                'max_y': max_y,
                'width': width,
                'height': height,
                'box': box
            })

        if not text_info:
            print("No valid text blocks detected")
            return None

        # Detect signatures
        print("Detecting signatures...")
        text_signatures = detect_signature_patterns(text_info, img.shape)
        visual_signatures = analyze_signature_regions(img, text_info)
        
        all_signatures = text_signatures + visual_signatures
        signature_blocks = {block['block']['text']: block for block in all_signatures}
        
        print(f"Found {len(all_signatures)} potential signatures")

        # Use HuggingFace for intelligent structure analysis
        hf_analysis = None
        if hf_processor:
            print("Analyzing table structure with HuggingFace...")
            hf_analysis = hf_processor.analyze_table_structure(text_info)
            print(f"HF Analysis: {hf_analysis.get('analysis', 'No analysis provided')}")
            
            # Build table based on HuggingFace analysis
            headers = hf_analysis.get('headers', [])
            
            # Create structured table data based on AI recommendations
            if hf_analysis.get('data_rows') and headers:
                structured_data = []
                signature_row_indices = []
                
                # Sort text_info by Y coordinate for row processing
                sorted_text_info = sorted(enumerate(text_info), key=lambda x: x[1]['center_y'])
                
                # Process each data row from HF analysis
                for row_info in hf_analysis['data_rows']:
                    row_data = [''] * len(headers)
                    
                    # Get blocks for this row and sort by X coordinate
                    row_blocks = []
                    for block_idx in row_info.get('blocks', []):
                        if block_idx < len(text_info):
                            row_blocks.append((block_idx, text_info[block_idx]))
                    
                    # Sort blocks by X coordinate for proper column assignment
                    row_blocks.sort(key=lambda x: x[1]['center_x'])
                    
                    # Assign blocks to columns based on X position
                    for i, (block_idx, block) in enumerate(row_blocks):
                        col_idx = min(i, len(headers) - 1)  # Don't exceed available columns
                        
                        if row_data[col_idx]:
                            row_data[col_idx] += ' ' + block['text']
                        else:
                            row_data[col_idx] = block['text']
                        
                        # Check for signatures
                        if block['text'] in signature_blocks:
                            if len(structured_data) not in signature_row_indices:
                                signature_row_indices.append(len(structured_data))
                    
                    # Mark special rows
                    if row_info.get('special_type') in ['signature', 'total', 'summary']:
                        if len(structured_data) not in signature_row_indices:
                            signature_row_indices.append(len(structured_data))
                    
                    # Only add non-empty rows
                    if any(cell.strip() for cell in row_data):
                        structured_data.append(row_data)
                
                # Create DataFrame with HuggingFace-structured data
                if structured_data:
                    df = pd.DataFrame(structured_data, columns=headers)
                else:
                    print("No structured data created, falling back to traditional method")
                    df, signature_row_indices = create_traditional_table(text_info, signature_blocks)
                
            else:
                # Fallback to original method if HF analysis is incomplete
                print("HuggingFace analysis incomplete, using traditional method")
                df, signature_row_indices = create_traditional_table(text_info, signature_blocks)
        else:
            # Traditional table creation
            df, signature_row_indices = create_traditional_table(text_info, signature_blocks)

        # Save outputs
        save_table_outputs(df, signature_row_indices, all_signatures, text_info, signature_blocks, output_dir)
        
        return df, {
            "table_data": {
                "headers": df.columns.tolist(),
                "rows": df.values.tolist(),
                "signature_rows": signature_row_indices
            },
            "signature_analysis": {
                "detected_signatures": [{
                    "text": sig['block']['text'],
                    "confidence": sig.get('confidence', sig.get('signature_score', 0)),
                    "type": sig['type'],
                    "position": {
                        "center_x": float(sig['block']['center_x']),
                        "center_y": float(sig['block']['center_y'])
                    }
                } for sig in all_signatures]
            },
            "hf_analysis": hf_analysis if hf_processor else None
        }
    
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def create_traditional_table(text_info, signature_blocks):
    """
    Traditional table creation method (fallback)
    """
    # Sort by vertical position
    text_info.sort(key=lambda x: x['center_y'])

    # Group into rows
    rows = []
    current_row = [text_info[0]]
    
    heights = [item['height'] for item in text_info]
    avg_height = sum(heights) / len(heights)
    row_height_threshold = avg_height * 0.8
    
    for i in range(1, len(text_info)):
        current_block = text_info[i]
        reference_block = current_row[0]
        
        y_overlap = min(current_block['max_y'], reference_block['max_y']) - max(current_block['min_y'], reference_block['min_y'])
        
        if y_overlap > 0 or abs(current_block['center_y'] - reference_block['center_y']) < row_height_threshold:
            current_row.append(current_block)
        else:
            current_row.sort(key=lambda x: x['center_x'])
            rows.append(current_row)
            current_row = [current_block]

    if current_row:
        current_row.sort(key=lambda x: x['center_x'])
        rows.append(current_row)

    # Determine column structure
    all_centers_x = [block['center_x'] for row in rows for block in row]
    
    column_centers = []
    if len(all_centers_x) > 10:
        sorted_x = sorted(all_centers_x)
        gaps = [(sorted_x[i+1] - sorted_x[i], i) for i in range(len(sorted_x)-1)]
        gaps.sort(reverse=True)
        
        num_columns = min(7, len(gaps) // 3 + 2)
        
        separators = sorted([sorted_x[gap[1]] for gap in gaps[:num_columns-1]])
        
        column_centers = [sorted_x[0] / 2]
        for i in range(len(separators)):
            mid_point = (separators[i] + (separators[i+1] if i+1 < len(separators) else sorted_x[-1])) / 2
            column_centers.append(mid_point)
    
    if not column_centers or len(column_centers) < 3:
        num_columns = max(len(row) for row in rows)
        
        min_x = min(block['min_x'] for row in rows for block in row)
        max_x = max(block['max_x'] for row in rows for block in row)
        
        column_width = (max_x - min_x) / num_columns
        column_centers = [min_x + column_width * (i + 0.5) for i in range(num_columns)]

    # Create table matrix
    table_matrix = []
    signature_row_indices = []
    
    for row_idx, row in enumerate(rows):
        row_data = [''] * len(column_centers)
        row_has_signature = False
        
        for block in row:
            distances = [abs(block['center_x'] - center) for center in column_centers]
            closest_col = distances.index(min(distances))
            
            if row_data[closest_col]:
                row_data[closest_col] += ' ' + block['text']
            else:
                row_data[closest_col] = block['text']
            
            if block['text'] in signature_blocks:
                row_has_signature = True
        
        table_matrix.append(row_data)
        if row_has_signature:
            signature_row_indices.append(row_idx)

    # Determine headers
    headers = []
    header_row_idx = 0
    
    if table_matrix and any(cell for cell in table_matrix[0] if cell.lower() in ['s. no', 's.no', 'sno', 'sl. no', 'serial no']):
        headers = table_matrix[0]
    else:
        non_empty_cols = sum(1 for cell in table_matrix[0] if cell.strip())
        if non_empty_cols >= min(3, len(table_matrix[0])):
            headers = table_matrix[0]
        else:
            headers = [f"Column {i+1}" for i in range(len(column_centers))]
            header_row_idx = -1

    headers = [h.strip() or f"Column {i+1}" for i, h in enumerate(headers)]
    
    if header_row_idx >= 0:
        data_rows = table_matrix[header_row_idx+1:]
        signature_row_indices = [idx - (header_row_idx + 1) for idx in signature_row_indices if idx > header_row_idx]
    else:
        data_rows = table_matrix

    # Create DataFrame
    max_cols = len(headers)
    normalized_rows = []
    final_signature_indices = []
    
    for row_idx, row in enumerate(data_rows):
        if not any(cell.strip() for cell in row):
            continue
            
        if len(row) < max_cols:
            row = row + [''] * (max_cols - len(row))
        elif len(row) > max_cols:
            row = row[:max_cols]
            
        normalized_rows.append(row)
        
        if row_idx in signature_row_indices:
            final_signature_indices.append(len(normalized_rows) - 1)

    if not normalized_rows:
        print("No valid table data extracted")
        return None, []

    df = pd.DataFrame(normalized_rows, columns=headers)
    return df, final_signature_indices

def save_table_outputs(df, signature_row_indices, all_signatures, text_info, signature_blocks, output_dir):
    """
    Save table outputs in various formats
    """
    # Save to CSV
    csv_path = os.path.join(output_dir, "extracted_table.csv")
    df.to_csv(csv_path, index=False)
    print(f"Table saved to CSV: {csv_path}")

    # Save to Excel with signature highlighting
    excel_path = os.path.join(output_dir, "extracted_table.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Table"
    
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    for sig_idx in signature_row_indices:
        excel_row = sig_idx + 2
        for col in range(1, len(df.columns) + 1):
            ws.cell(row=excel_row, column=col).fill = green_fill
    
    wb.save(excel_path)
    print(f"Table saved to Excel with signature highlighting: {excel_path}")

    # Save to TXT
    txt_path = os.path.join(output_dir, "extracted_table.txt")
    with open(txt_path, 'w', encoding='utf-8') as txt_file:
        txt_file.write("\t".join(df.columns) + "\n")
        txt_file.write("-" * 80 + "\n")
        
        for idx, (_, row) in enumerate(df.iterrows()):
            prefix = "[SIGNATURE] " if idx in signature_row_indices else ""
            txt_file.write(prefix + "\t".join(str(cell) for cell in row) + "\n")
    
    print(f"Table saved to TXT: {txt_path}")

    # Save JSON with analysis
    json_path = os.path.join(output_dir, "extracted_table.json")
    json_data = {
        "table_data": {
            "headers": df.columns.tolist(),
            "rows": df.values.tolist(),
            "signature_rows": signature_row_indices
        },
        "signature_analysis": {
            "detected_signatures": [{
                "text": sig['block']['text'],
                "confidence": sig.get('confidence', sig.get('signature_score', 0)),
                "type": sig['type'],
                "position": {
                    "center_x": float(sig['block']['center_x']),
                    "center_y": float(sig['block']['center_y'])
                }
            } for sig in all_signatures]
        },
        "raw_data": {
            "text_blocks": [{
                "text": item["text"],
                "confidence": float(item["confidence"]),
                "position": {
                    "center_x": float(item["center_x"]),
                    "center_y": float(item["center_y"]),
                    "box": [[float(p[0]), float(p[1])] for p in item["box"]]
                },
                "is_signature": item["text"] in signature_blocks
            } for item in text_info]
        }
    }
    
    with open(json_path, 'w', encoding='utf-8') as json_file:
        json.dump(json_data, json_file, indent=2, ensure_ascii=False)
    
    print(f"Table saved to JSON: {json_path}")
    print("\nExtracted table content:")
    print(df)
    
    if signature_row_indices:
        print(f"\nRows with detected signatures: {signature_row_indices}")

def main():
    image_path = "/home/talgotram/Repos/ioclOCR/input/images/page_2.jpg"
    output_dir = "/home/talgotram/Repos/ioclOCR/output/table_output"
    
    # Set your OpenAI API key here or set OPENAI_API_KEY environment variable
    openai_api_key = os.getenv('OPENAI_API_KEY')
    
    try:
        result = enhanced_table_extraction(
            image_path, 
            output_dir, 
        )
        
        if result:
            df, analysis_data = result
            print("\n" + "="*50)
            print("EXTRACTION COMPLETE")
            print("="*50)
            print(f"Extracted {len(df)} rows with {len(df.columns)} columns")
            if analysis_data.get('llm_analysis'):
                print("LLM-enhanced structure analysis was used")
            else:
                print("Traditional extraction method was used")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()