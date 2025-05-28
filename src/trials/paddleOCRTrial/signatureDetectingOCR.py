import cv2
import pandas as pd
import os
import json
import numpy as np
import re
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import time
from dotenv import load_dotenv

load_dotenv()

def detect_signature_patterns(text_blocks, img_shape):
    signature_indicators = []
    
    signature_keywords = [
        'signature', 'signed', 'sign', 'name', 'initial', 'initials',
        'authorized', 'approved', 'verified', 'confirm', 'acknowledge',
        'witness', 'attest', 'certify', 'endorse'
    ]
    
    signature_patterns = [
        r'\b[A-Z][a-z]+ [A-Z][a-z]+\b',
        r'\b[A-Z]\.[A-Z]\. [A-Z][a-z]+\b',
        r'\b[A-Z]{2,4}\b',
        r'\/s\/ .+',
        r'_+\s*$',
        r'-{3,}',
    ]
    
    for block in text_blocks:
        text = block['text'].strip()
        is_signature = False
        confidence_score = 0
        
        if any(keyword in text.lower() for keyword in signature_keywords):
            is_signature = True
            confidence_score += 0.4
        
        for pattern in signature_patterns:
            if re.search(pattern, text):
                is_signature = True
                confidence_score += 0.3
        
        if block['confidence'] < 0.7:
            confidence_score += 0.2
        
        aspect_ratio = block['width'] / block['height'] if block['height'] > 0 else 0
        if aspect_ratio > 3 or aspect_ratio < 0.5:
            confidence_score += 0.1
        
        if len(text.split()) <= 3 and len(text) > 2:
            confidence_score += 0.2
        
        if re.match(r'^[A-Z][a-z]*\.?\s*[A-Z][a-z]*\.?$', text):  # Name patterns
            confidence_score += 0.3
        
        if confidence_score >= 0.5:
            signature_indicators.append({
                'block': block,
                'confidence': confidence_score,
                'type': 'text_signature'
            })
    
    return signature_indicators

def analyze_signature_regions(img, text_blocks):
    signature_regions = []
    
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    for block in text_blocks:
        x1, y1 = int(block['min_x']), int(block['min_y'])
        x2, y2 = int(block['max_x']), int(block['max_y'])
        
        x1 = max(0, x1 - 10)
        y1 = max(0, y1 - 10)
        x2 = min(gray.shape[1], x2 + 10)
        y2 = min(gray.shape[0], y2 + 10)
        
        if x2 > x1 and y2 > y1:
            region = gray[y1:y2, x1:x2]
            
            if region.size > 0:
                variance = np.var(region)
                
                edges = cv2.Canny(region, 50, 150)
                edge_density = np.sum(edges > 0) / edges.size
                
                signature_score = (variance / 1000) + (edge_density * 2)
                
                if signature_score > 0.3:  # Threshold for signature detection
                    signature_regions.append({
                        'block': block,
                        'signature_score': signature_score,
                        'type': 'visual_signature'
                    })
    
    return signature_regions

class MistralTableEnhancer:
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = "https://api.mistral.ai/v1/chat/completions"
        self.headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
    
    def enhance_table_structure(self, df, signature_indices=None, max_retries=3):
        """
        Use mistral API to analyze and improve table structure
        """
        for attempt in range(max_retries):
            try:
                # Prepare the table data for Mistral
                table_preview = df.head(10).to_string(index=False)
                
                # Get more comprehensive table data
                table_preview = df.to_string(index=True, max_rows=15)
                sample_data = []
                for i, row in df.head(10).iterrows():
                    sample_data.append([str(cell).strip() for cell in row.values])
                
                prompt = f"""
                    You are a data structure expert. Analyze this OCR-extracted table and improve its organization.

                    CURRENT TABLE STRUCTURE:
                    Headers: {list(df.columns)}
                    Number of columns: {len(df.columns)}
                    Number of rows: {len(df)}

                    RAW TABLE DATA:
                    {table_preview}

                    SAMPLE ROWS (first 10):
                    {sample_data}

                    ANALYSIS REQUIREMENTS:
                    1. Look at the actual data content in each column to determine what it represents
                    2. Identify if generic headers like "Column 1", "Column 2" should be replaced with meaningful names
                    3. Reorganize the data such that each column should have data which only belongs to that column
                    4. Ensure ALL {len(df.columns)} columns get proper headers - do not skip any
                    5. Look for common table patterns: ID/Serial numbers, Names, Dates, Amounts, Status, etc.
                    - Extract as a nested array structure with headers
                    - Maintain column headers and row labels
                    - Preserve all cell values with their exact formatting
                    - Handle merged cells appropriately

                    STRICT RULES:
                    - You MUST provide {len(df.columns)} or more improved headers
                    - If you can't determine a column's purpose, use descriptive names like "Field_1", "Data_2", etc.
                    - Headers should be concise but descriptive
                    - No special characters in headers, use underscores instead of spaces
                    - Check if data appears misaligned (content that should be in one column appears in another)
                    - Data SHOULD NOT be in the incorrect column, check multiple times before giving final output

                    Return ONLY this JSON structure:
                    OUTPUT FORMAT:
                    Return ONLY valid JSON with this structure:
                    {{
                        "document_type": "detected document type (e.g., invoice, form, report)",
                        "page_metadata": {{
                            "page_number": "detected page number if present",
                            "header": "header text if present",
                            "footer": "footer text if present"
                        }},
                        "sections": [
                            {{
                                "section_type": "text|table|form|chart",
                                "section_title": "section heading if present",
                                "content": "appropriate content structure based on section type"
                            }}
                        ],
                        "tables": [
                            {{
                                "table_title": "title if present",
                                "headers": ["header1", "header2", ...],
                                "data": [
                                    ["row1col1", "row1col2", ...],
                                    ["row2col1", "row2col2", ...]
                                ]
                            }}
                        ],
                        "key_value_pairs": {{
                            "key1": "value1",
                            "key2": "value2"
                        }}
                    }}

                    Focus on practical improvements that make the table more professional and readable.
                    Make sure to use proper JSON escaping for special characters and ensure the output is valid JSON.
                    If certain elements don't exist, include them as empty arrays or objects rather than omitting them."""

                payload = {
                    "messages": [
                        {"role": "user", "content": prompt}
                    ],
                    "model": "mistral-large-latest",
                    "stream": False,
                    "temperature": 0.1,
                    "response_format": {
                        "type": "text",
                        "json_schema": {{
                            "document_type": "detected document type (e.g., invoice, form, report)",
                            "page_metadata": {{
                                "page_number": "detected page number if present",
                                "header": "header text if present",
                                "footer": "footer text if present"
                            }},
                            "sections": [
                                {{
                                    "section_type": "text|table|form|chart",
                                    "section_title": "section heading if present",
                                    "content": "appropriate content structure based on section type"
                                }}
                            ],
                            "tables": [
                                {{
                                    "table_title": "title if present",
                                    "headers": ["header1", "header2", ...],
                                    "data": [
                                        ["row1col1", "row1col2", ...],
                                        ["row2col1", "row2col2", ...]
                                    ]
                                }}
                            ],
                            "key_value_pairs": {{
                                "key1": "value1",
                                "key2": "value2"
                            }}
                        }}
                    },
                }

                response = requests.post(
                    self.base_url,
                    headers=self.headers,
                    json=payload,
                    timeout=30
                )

                if response.status_code == 200:
                    result = response.json()
                    mistral_response = result['choices']
                    print(mistral_response)
                    
                    return
                    # Try to extract JSON from the response
                    try:
                        # Look for JSON block in the response
                        import re
                        json_match = re.search(r'```json\n(.*?)\n```', mistral_response, re.DOTALL)
                        if json_match:
                            json_str = json_match.group(1)
                        else:
                            # Try to find JSON without code blocks
                            json_match = re.search(r'\{.*\}', mistral_response, re.DOTALL)
                            if json_match:
                                json_str = json_match.group(0)
                            else:
                                raise ValueError("No JSON found in response")
                        
                        enhancement_data = json.loads(json_str)
                        return self.apply_enhancements(df, enhancement_data, signature_indices)
                    
                    except (json.JSONDecodeError, ValueError) as e:
                        print(f"Failed to parse JSON from mistral response: {e}")
                        print(f"Raw response: {mistral_response}")
                        return df, {"analysis": "Failed to parse mistral response"}
                
                else:
                    print(f"mistral API error: {response.status_code} - {response.text}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)  # Exponential backoff
                        continue
                    else:
                        return df, {"analysis": "Failed to get response from mistral API"}

            except requests.exceptions.RequestException as e:
                print(f"Request failed (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
                else:
                    return df, {"analysis": f"Network error: {str(e)}"}
        
        return df, {"analysis": "All retry attempts failed"}
    
    def apply_enhancements(self, df, enhancement_data, signature_indices):
        """
        Apply the enhancements suggested by mistral
        """
        try:
            enhanced_df = df.copy()
            
            # Apply column renaming
            if 'improved_headers' in enhancement_data:
                new_headers = enhancement_data['improved_headers'][:len(df.columns)]
                if len(new_headers) == len(df.columns):
                    enhanced_df.columns = new_headers
            
            elif 'column_mapping' in enhancement_data:
                enhanced_df = enhanced_df.rename(columns=enhancement_data['column_mapping'])
            
            # Apply data corrections
            if 'data_corrections' in enhancement_data:
                for correction in enhancement_data['data_corrections']:
                    try:
                        row_idx = correction.get('row_index')
                        column = correction.get('column')
                        value = correction.get('corrected_value')
                        
                        if row_idx is not None and column in enhanced_df.columns and row_idx < len(enhanced_df):
                            enhanced_df.at[row_idx, column] = value
                    except Exception as e:
                        print(f"Failed to apply correction: {e}")
            
            return enhanced_df, enhancement_data
            
        except Exception as e:
            print(f"Error applying enhancements: {e}")
            return df, enhancement_data

# Function to add mistral enhancement to your existing workflow
def enhance_with_mistral(df, signature_indices, mistral_api_key, output_dir):
    """
    Enhance the extracted table using Mistral API
    """
    if not mistral_api_key:
        print("No mistral API key provided, skipping enhancement")
        return df, {}
    
    print("Enhancing table structure with mistral API...")
    
    mistral_enhancer = MistralTableEnhancer(mistral_api_key)
    enhanced_df, enhancement_info = mistral_enhancer.enhance_table_structure(df, signature_indices)
    
    # Save the enhanced table
    if not enhanced_df.equals(df):
        print("mistral suggested improvements applied!")
        
        # Save enhanced versions
        enhanced_csv_path = os.path.join(output_dir, "enhanced_table.csv")
        enhanced_df.to_csv(enhanced_csv_path, index=False)
        print(f"Enhanced table saved to: {enhanced_csv_path}")
        
        enhanced_excel_path = os.path.join(output_dir, "enhanced_table.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Enhanced Table"
        
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        for r in dataframe_to_rows(enhanced_df, index=False, header=True):
            ws.append(r)
        
        # Highlight signature rows
        for sig_idx in signature_indices:
            excel_row = sig_idx + 2
            for col in range(1, len(enhanced_df.columns) + 1):
                ws.cell(row=excel_row, column=col).fill = green_fill
        
        wb.save(enhanced_excel_path)
        print(f"Enhanced Excel saved to: {enhanced_excel_path}")
        
        # Save enhancement analysis
        enhancement_path = os.path.join(output_dir, "mistral_analysis.json")
        with open(enhancement_path, 'w', encoding='utf-8') as f:
            json.dump(enhancement_info, f, indent=2, ensure_ascii=False)
        print(f"Mistral analysis saved to: {enhancement_path}")
        
        print(f"\nMistral Analysis: {enhancement_info.get('analysis', 'No analysis provided')}")
        
    else:
        print("No improvements suggested by Mistral")
    
    return enhanced_df, enhancement_info

def improved_table_extraction(image_path, mistral_api_key, output_dir='./output'):
    os.makedirs(output_dir, exist_ok=True)

    ocr = PaddleOCR(
        use_angle_cls=True, 
        lang='en',
        use_gpu=True,
        show_log=False,
        draw_img_save_dir=None,
        vis_font_path=None
    )

    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f"Could not load image from {image_path}")
    
    try:
        result = ocr.ocr(img, cls=True)
        
        if not result or not result[0]:
            print("No text detected in the image")
            return None

        text_info = []
        for line in result[0]:
            box = line[0]
            text = line[1][0].strip()
            confidence = line[1][1]

            if not text:
                continue

            x_coords = [p[0] for p in box]
            y_coords = [p[1] for p in box]
            
            min_x = min(x_coords)
            max_x = max(x_coords)
            min_y = min(y_coords)
            max_y = max(y_coords)
            
            width = max_x - min_x
            height = max_y - min_y
            
            center_x = (min_x + max_x) / 2
            center_y = (min_y + max_y) / 2
            
            text_info.append({
                'text': text,
                'confidence': confidence,
                'center_x': center_x,
                'center_y': center_y,
                'min_x': min_x,
                'max_x': max_x,
                'min_y': min_y,
                'max_y': max_y,
                'width': width,
                'height': height,
                'box': box
            })

        if not text_info:
            print("No valid text blocks detected")
            return None

        print("Detecting signatures...")
        text_signatures = detect_signature_patterns(text_info, img.shape)
        visual_signatures = analyze_signature_regions(img, text_info)
        
        all_signatures = text_signatures + visual_signatures
        signature_blocks = {block['block']['text']: block for block in all_signatures}
        
        print(f"Found {len(all_signatures)} potential signatures")
        for sig in all_signatures:
            print(f"  - '{sig['block']['text']}' (confidence: {sig.get('confidence', sig.get('signature_score', 0)):.2f})")

        text_info.sort(key=lambda x: x['center_y'])

        rows = []
        current_row = [text_info[0]]
        
        heights = [item['height'] for item in text_info]
        avg_height = sum(heights) / len(heights)
        row_height_threshold = avg_height * 0.8
        
        for i in range(1, len(text_info)):
            current_block = text_info[i]
            reference_block = current_row[0]
            
            y_overlap = min(current_block['max_y'], reference_block['max_y']) - max(current_block['min_y'], reference_block['min_y'])
            
            if y_overlap > 0 or abs(current_block['center_y'] - reference_block['center_y']) < row_height_threshold:
                current_row.append(current_block)
            else:
                current_row.sort(key=lambda x: x['center_x'])
                rows.append(current_row)
                current_row = [current_block]

        if current_row:
            current_row.sort(key=lambda x: x['center_x'])
            rows.append(current_row)

        all_centers_x = [block['center_x'] for row in rows for block in row]
        
        column_centers = []
        if len(all_centers_x) > 10:
            sorted_x = sorted(all_centers_x)
            gaps = [(sorted_x[i+1] - sorted_x[i], i) for i in range(len(sorted_x)-1)]
            gaps.sort(reverse=True)
            
            num_columns = min(7, len(gaps) // 3 + 2)
            
            separators = sorted([sorted_x[gap[1]] for gap in gaps[:num_columns-1]])
            
            column_centers = [sorted_x[0] / 2]  # Start
            for i in range(len(separators)):
                mid_point = (separators[i] + (separators[i+1] if i+1 < len(separators) else sorted_x[-1])) / 2
                column_centers.append(mid_point)
        
        if not column_centers or len(column_centers) < 3:
            num_columns = max(len(row) for row in rows)
            
            min_x = min(block['min_x'] for row in rows for block in row)
            max_x = max(block['max_x'] for row in rows for block in row)
            
            column_width = (max_x - min_x) / num_columns
            column_centers = [min_x + column_width * (i + 0.5) for i in range(num_columns)]

        table_matrix = []
        signature_row_indices = []
        
        for row_idx, row in enumerate(rows):
            row_data = [''] * len(column_centers)
            row_has_signature = False
            
            for block in row:
                distances = [abs(block['center_x'] - center) for center in column_centers]
                closest_col = distances.index(min(distances))
                
                if row_data[closest_col]:
                    row_data[closest_col] += ' ' + block['text']
                else:
                    row_data[closest_col] = block['text']
                
                if block['text'] in signature_blocks:
                    row_has_signature = True
            
            table_matrix.append(row_data)
            if row_has_signature:
                signature_row_indices.append(row_idx)

        
        headers = []
        header_row_idx = 0
        
        if table_matrix and any(cell for cell in table_matrix[0] if cell.lower() in ['s. no', 's.no', 'sno', 'sl. no', 'serial no']):
            headers = table_matrix[0]
        else:
            non_empty_cols = sum(1 for cell in table_matrix[0] if cell.strip())
            if non_empty_cols >= min(3, len(table_matrix[0])):
                headers = table_matrix[0]
            else:
                headers = [f"Column {i+1}" for i in range(len(column_centers))]
                header_row_idx = -1  # No header row to skip
        
        headers = [h.strip() or f"Column {i+1}" for i, h in enumerate(headers)]
        
        if header_row_idx >= 0:
            data_rows = table_matrix[header_row_idx+1:]
            signature_row_indices = [idx - (header_row_idx + 1) for idx in signature_row_indices if idx > header_row_idx]
        else:
            data_rows = table_matrix
            
        max_cols = len(headers)
        normalized_rows = []
        final_signature_indices = []
        
        for row_idx, row in enumerate(data_rows):
            if not any(cell.strip() for cell in row):
                continue
                
            if len(row) < max_cols:
                row = row + [''] * (max_cols - len(row))
            elif len(row) > max_cols:
                row = row[:max_cols]
                
            normalized_rows.append(row)
            
            if row_idx in signature_row_indices:
                final_signature_indices.append(len(normalized_rows) - 1)

        if not normalized_rows:
            print("No valid table data extracted")
            return None

        df = pd.DataFrame(normalized_rows, columns=headers)
        
        cleaned_df = df.copy()
        cleaned_signature_indices = final_signature_indices.copy()
        
        i = 0
        while i < len(cleaned_df) - 1:
            current_row = cleaned_df.iloc[i]
            next_row = cleaned_df.iloc[i+1]
            
            current_non_empty = current_row.astype(str).str.strip().ne('').sum()
            next_non_empty = next_row.astype(str).str.strip().ne('').sum()
            
            if next_non_empty <= max_cols // 2 and current_non_empty > next_non_empty:
                for col in cleaned_df.columns:
                    if pd.notna(next_row[col]) and next_row[col].strip():
                        if pd.notna(current_row[col]) and current_row[col].strip():
                            cleaned_df.at[i, col] = f"{current_row[col]} {next_row[col]}"
                        else:
                            cleaned_df.at[i, col] = next_row[col]
    
                if i+1 in cleaned_signature_indices:
                    cleaned_signature_indices.remove(i+1)
                    if i not in cleaned_signature_indices:
                        cleaned_signature_indices.append(i)
                
                cleaned_df = cleaned_df.drop(i+1)
                cleaned_df = cleaned_df.reset_index(drop=True)

                cleaned_signature_indices = [idx if idx <= i else idx - 1 for idx in cleaned_signature_indices]
            else:
                i += 1
        
        if 'S. No' in cleaned_df.columns or any(col for col in cleaned_df.columns if 'no' in col.lower()):
            sno_col = next((col for col in cleaned_df.columns if 'no' in col.lower()), cleaned_df.columns[0])
            
            cleaned_df[sno_col] = pd.to_numeric(cleaned_df[sno_col], errors='coerce')
            
            mask = cleaned_df[sno_col].isna()
            if mask.any():
                valid_sns = cleaned_df.loc[~mask, sno_col].dropna()
                if not valid_sns.empty:
                    last_sn = valid_sns.iloc[-1]
                    counter = last_sn + 1
                    
                    for idx in cleaned_df.index[mask]:
                        cleaned_df.at[idx, sno_col] = counter
                        counter += 1

        csv_path = os.path.join(output_dir, "extracted_table.csv")
        cleaned_df.to_csv(csv_path, index=False)
        print(f"Table saved to CSV: {csv_path}")

        excel_path = os.path.join(output_dir, "extracted_table.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Table"

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for r in dataframe_to_rows(cleaned_df, index=False, header=True):
            ws.append(r)

        for sig_idx in cleaned_signature_indices:
            excel_row = sig_idx + 2  # +1 for header, +1 for 1-indexed
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).fill = green_fill
        
        wb.save(excel_path)
        print(f"Table saved to Excel with signature highlighting: {excel_path}")
        print(f"Highlighted {len(cleaned_signature_indices)} rows with detected signatures")

        txt_path = os.path.join(output_dir, "extracted_table.txt")
        with open(txt_path, 'w', encoding='utf-8') as txt_file:
            txt_file.write("\t".join(cleaned_df.columns) + "\n")
            txt_file.write("-" * 80 + "\n")
            
            for idx, (_, row) in enumerate(cleaned_df.iterrows()):
                prefix = "[SIGNATURE] " if idx in cleaned_signature_indices else ""
                txt_file.write(prefix + "\t".join(str(cell) for cell in row) + "\n")
        
        print(f"Table saved to TXT: {txt_path}")

        json_path = os.path.join(output_dir, "extracted_table.json")
        json_data = {
            "table_data": {
                "headers": cleaned_df.columns.tolist(),
                "rows": cleaned_df.values.tolist(),
                "signature_rows": cleaned_signature_indices
            },
            "signature_analysis": {
                "detected_signatures": [{
                    "text": sig['block']['text'],
                    "confidence": sig.get('confidence', sig.get('signature_score', 0)),
                    "type": sig['type'],
                    "position": {
                        "center_x": float(sig['block']['center_x']),
                        "center_y": float(sig['block']['center_y'])
                    }
                } for sig in all_signatures]
            },
            "raw_data": {
                "text_blocks": [{
                    "text": item["text"],
                    "confidence": float(item["confidence"]),
                    "position": {
                        "center_x": float(item["center_x"]),
                        "center_y": float(item["center_y"]),
                        "box": [[float(p[0]), float(p[1])] for p in item["box"]]
                    },
                    "is_signature": item["text"] in signature_blocks
                } for item in text_info]
            }
        }
        
        with open(json_path, 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)
        
        print(f"Table saved to JSON: {json_path}")

        print("\nExtracted table content:")
        #print(cleaned_df)

        api_key = mistral_api_key
        enhanced_df, mistral_analysis = enhance_with_mistral(cleaned_df, cleaned_signature_indices, api_key, output_dir)

        return enhanced_df, json_data
    
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    image_path = "/home/talgotram/Repos/ioclOCR/input/images/page_1.jpg"
    output_dir = "/home/talgotram/Repos/ioclOCR/output/table_output_1"
    
    mistral_api_key = os.getenv('MISTRAL_API_KEY')

    try:
        improved_table_extraction(image_path, mistral_api_key, output_dir)
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()