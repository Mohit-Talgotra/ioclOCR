import cv2
import pandas as pd
import os
import json
import numpy as np
import re
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import time
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv

load_dotenv()

class HuggingFaceTableProcessor: 
    def __init__(self, api_token: Optional[str] = None):
        """
        Initialize the HuggingFace processor
        
        Args:
            api_token: HuggingFace API token (required for API access)
        """
        if not api_token:
            raise ValueError("HuggingFace API token is required. Get one from https://huggingface.co/settings/tokens")
            
        self.api_token = api_token
        self.model = "microsoft/DialoGPT-medium"
        
        self.headers = {
            "Authorization": f"Bearer {api_token}",
            "Content-Type": "application/json"
        }
        
        self.base_url = "https://api-inference.huggingface.co/models"
        
        self._test_api_connection()
        
    def _test_api_connection(self):
        test_url = "https://api-inference.huggingface.co/models/microsoft/DialoGPT-medium"
        test_payload = {"inputs": "test"}
        
        try:
            response = requests.post(test_url, headers=self.headers, json=test_payload, timeout=10)
            if response.status_code == 401:
                raise ValueError("Invalid HuggingFace API token. Please check your token at https://huggingface.co/settings/tokens")
            elif response.status_code == 403:
                raise ValueError("API token doesn't have required permissions")
        except requests.exceptions.RequestException as e:
            print(f"Warning: Could not verify API connection: {e}")
        
    def _make_api_call(self, payload: dict, max_retries: int = 3) -> dict:
        url = f"{self.base_url}/{self.model}"
        
        for attempt in range(max_retries):
            try:
                response = requests.post(url, headers=self.headers, json=payload, timeout=30)
                
                if response.status_code == 401:
                    raise ValueError("Invalid API credentials. Check your HuggingFace token.")
                elif response.status_code == 403:
                    raise ValueError("API access forbidden. Check token permissions.")
                elif response.status_code == 503:
                    wait_time = 20 * (attempt + 1)
                    print(f"Model loading, waiting {wait_time} seconds...")
                    time.sleep(wait_time)
                    continue
                elif response.status_code == 200:
                    return response.json()
                else:
                    print(f"API call failed with status {response.status_code}: {response.text}")
                    
            except requests.exceptions.RequestException as e:
                print(f"Request failed (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(5)
                    
        return None
    
    def analyze_table_structure(self, raw_text_blocks: List[Dict]) -> Dict[str, Any]:
        text_content = []
        for i, block in enumerate(raw_text_blocks[:30]):  # Reduced to avoid token limits
            text_content.append(f"{i}: '{block['text']}' (y:{int(block['center_y'])})")
        
        analysis_prompt = f"""Text blocks from a table:
{chr(10).join(text_content[:20])}

Identify headers and structure. List header block numbers and suggest column names."""

        try:
            payload = {
                "inputs": analysis_prompt,
                "parameters": {
                    "max_new_tokens": 150,
                    "temperature": 0.1,
                    "do_sample": False,
                    "return_full_text": False
                }
            }
            
            result = self._make_api_call(payload)
            
            if result:
                return self._parse_hf_response(result, raw_text_blocks)
            else:
                print("HuggingFace analysis failed, using fallback")
                return self._fallback_analysis(raw_text_blocks)
                
        except Exception as e:
            print(f"HuggingFace analysis failed: {e}")
            return self._fallback_analysis(raw_text_blocks)
    
    def _parse_hf_response(self, hf_result: dict, raw_text_blocks: List[Dict]) -> Dict[str, Any]:
        try:
            if isinstance(hf_result, list) and len(hf_result) > 0:
                text_response = hf_result[0].get('generated_text', '')
            elif isinstance(hf_result, dict):
                text_response = hf_result.get('generated_text', str(hf_result))
            else:
                text_response = str(hf_result)
            
            headers = []
            header_blocks = []
            
            numbers = re.findall(r'\b\d+\b', text_response)
            if numbers:
                header_blocks = [int(num) for num in numbers[:6] if int(num) < len(raw_text_blocks)]
            
            if header_blocks:
                headers = [raw_text_blocks[i]['text'] for i in header_blocks]
            else:
                y_positions = sorted(set(block['center_y'] for block in raw_text_blocks))
                if y_positions:
                    first_row_y = y_positions[0]
                    headers = [block['text'] for block in raw_text_blocks 
                             if abs(block['center_y'] - first_row_y) < 20][:6]
            
            if not headers:
                headers = [f"Column_{i+1}" for i in range(5)]
            
            return {
                "headers": headers,
                "header_blocks": header_blocks,
                "data_rows": [],
                "analysis": f"HuggingFace analysis completed. Found {len(headers)} headers.",
                "raw_response": text_response[:200]
            }
            
        except Exception as e:
            print(f"Error parsing HuggingFace response: {e}")
            return self._fallback_analysis(raw_text_blocks)
    
    def _fallback_analysis(self, raw_text_blocks: List[Dict]) -> Dict[str, Any]:
        headers = []
        header_blocks = []
        
        y_positions = sorted(set(block['center_y'] for block in raw_text_blocks))
        first_row_y = y_positions[0] if y_positions else 0
        
        header_keywords = ['name', 'date', 'amount', 'description', 'type', 'status', 'id', 'number', 'total']
        
        for i, block in enumerate(raw_text_blocks):
            if abs(block['center_y'] - first_row_y) < 25:
                text_lower = block['text'].lower().strip()
                is_header = (
                    any(keyword in text_lower for keyword in header_keywords) or
                    len(text_lower.split()) <= 3 or
                    text_lower.replace('.', '').replace(' ', '').isalpha()
                )
                
                if is_header:
                    headers.append(block['text'])
                    header_blocks.append(i)
        
        if not headers:
            num_cols = len([block for block in raw_text_blocks 
                           if abs(block['center_y'] - first_row_y) < 25])
            headers = [f"Column_{i+1}" for i in range(min(num_cols, 6))]
        
        return {
            "headers": headers,
            "header_blocks": header_blocks,
            "data_rows": [],
            "analysis": "Fallback analysis used"
        }

def detect_signature_patterns(text_blocks, img_shape):
    signature_indicators = []
    
    signature_keywords = [
        'signature', 'signed', 'sign', 'name', 'initial', 'initials',
        'authorized', 'approved', 'verified', 'confirm', 'acknowledge',
        'witness', 'attest', 'certify', 'endorse'
    ]
    
    signature_patterns = [
        r'\b[A-Z][a-z]+ [A-Z][a-z]+\b',
        r'\b[A-Z]\.[A-Z]\. [A-Z][a-z]+\b',
        r'\b[A-Z]{2,4}\b',
        r'\/s\/ .+',
        r'_+\s*$',
        r'-{3,}',
    ]
    
    for block in text_blocks:
        text = block['text'].strip()
        confidence_score = 0
        
        if any(keyword in text.lower() for keyword in signature_keywords):
            confidence_score += 0.4
        
        for pattern in signature_patterns:
            if re.search(pattern, text):
                confidence_score += 0.3
        
        if block['confidence'] < 0.7:
            confidence_score += 0.2
        
        aspect_ratio = block['width'] / block['height'] if block['height'] > 0 else 0
        if aspect_ratio > 3 or aspect_ratio < 0.5:
            confidence_score += 0.1
        
        if len(text.split()) <= 3 and len(text) > 2:
            confidence_score += 0.2
        
        if re.match(r'^[A-Z][a-z]*\.?\s*[A-Z][a-z]*\.?$', text):
            confidence_score += 0.3
        
        if confidence_score >= 0.5:
            signature_indicators.append({
                'block': block,
                'confidence': confidence_score,
                'type': 'text_signature'
            })
    
    return signature_indicators

def analyze_signature_regions(img, text_blocks):
    signature_regions = []
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    for block in text_blocks:
        x1, y1 = int(block['min_x']), int(block['min_y'])
        x2, y2 = int(block['max_x']), int(block['max_y'])
        
        x1 = max(0, x1 - 10)
        y1 = max(0, y1 - 10)
        x2 = min(gray.shape[1], x2 + 10)
        y2 = min(gray.shape[0], y2 + 10)
        
        if x2 > x1 and y2 > y1:
            region = gray[y1:y2, x1:x2]
            
            if region.size > 0:
                variance = np.var(region)
                edges = cv2.Canny(region, 50, 150)
                edge_density = np.sum(edges > 0) / edges.size
                signature_score = (variance / 1000) + (edge_density * 2)
                
                if signature_score > 0.3:
                    signature_regions.append({
                        'block': block,
                        'signature_score': signature_score,
                        'type': 'visual_signature'
                    })
    
    return signature_regions

def enhanced_table_extraction(image_path, output_dir='./output', use_ai=True, hf_api_token=None):
    os.makedirs(output_dir, exist_ok=True)

    hf_processor = None
    if use_ai and hf_api_token:
        try:
            hf_processor = HuggingFaceTableProcessor(api_token=hf_api_token)
            print("HuggingFace processor initialized successfully")
        except Exception as e:
            print(f"Failed to initialize HuggingFace processor: {e}")
            print("Continuing with traditional extraction...")

    ocr = PaddleOCR(
        use_angle_cls=True, 
        lang='en',
        use_gpu=False,
        show_log=False
    )

    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f"Could not load image from {image_path}")
    
    result = ocr.ocr(img, cls=True)
    
    if not result or not result[0]:
        print("No text detected in the image")
        return None

    text_info = []
    for line in result[0]:
        box = line[0]
        text = line[1][0].strip()
        confidence = line[1][1]

        if not text:
            continue

        x_coords = [p[0] for p in box]
        y_coords = [p[1] for p in box]
        
        min_x = min(x_coords)
        max_x = max(x_coords)
        min_y = min(y_coords)
        max_y = max(y_coords)
        
        width = max_x - min_x
        height = max_y - min_y
        
        center_x = (min_x + max_x) / 2
        center_y = (min_y + max_y) / 2
        
        text_info.append({
            'text': text,
            'confidence': confidence,
            'center_x': center_x,
            'center_y': center_y,
            'min_x': min_x,
            'max_x': max_x,
            'min_y': min_y,
            'max_y': max_y,
            'width': width,
            'height': height,
            'box': box
        })

    if not text_info:
        print("No valid text blocks detected")
        return None

    print("Detecting signatures...")
    text_signatures = detect_signature_patterns(text_info, img.shape)
    visual_signatures = analyze_signature_regions(img, text_info)
    
    all_signatures = text_signatures + visual_signatures
    signature_blocks = {block['block']['text']: block for block in all_signatures}
    
    print(f"Found {len(all_signatures)} potential signatures")

    if hf_processor:
        print("Analyzing table structure with HuggingFace...")
        hf_analysis = hf_processor.analyze_table_structure(text_info)
        print(f"HF Analysis: {hf_analysis.get('analysis', 'No analysis provided')}")
        
        if hf_analysis.get('headers'):
            df, signature_row_indices = create_hf_based_table(text_info, hf_analysis, signature_blocks)
        else:
            df, signature_row_indices = create_traditional_table(text_info, signature_blocks)
    else:
        df, signature_row_indices = create_traditional_table(text_info, signature_blocks)

    save_table_outputs(df, signature_row_indices, all_signatures, text_info, signature_blocks, output_dir)
    
    return df, {
        "table_data": {
            "headers": df.columns.tolist(),
            "rows": df.values.tolist(),
            "signature_rows": signature_row_indices
        },
        "signature_analysis": {
            "detected_signatures": [{
                "text": sig['block']['text'],
                "confidence": sig.get('confidence', sig.get('signature_score', 0)),
                "type": sig['type'],
                "position": {
                    "center_x": float(sig['block']['center_x']),
                    "center_y": float(sig['block']['center_y'])
                }
            } for sig in all_signatures]
        }
    }

def create_hf_based_table(text_info, hf_analysis, signature_blocks):
    headers = hf_analysis['headers']
    
    sorted_blocks = sorted(enumerate(text_info), key=lambda x: x[1]['center_y'])
    
    header_block_indices = set(hf_analysis.get('header_blocks', []))
    data_blocks = [(i, block) for i, block in sorted_blocks if i not in header_block_indices]
    
    rows = []
    signature_row_indices = []
    
    if data_blocks:
        current_row = [data_blocks[0]]
        row_threshold = 30
        
        for i in range(1, len(data_blocks)):
            current_block = data_blocks[i]
            reference_block = current_row[0]
            
            if abs(current_block[1]['center_y'] - reference_block[1]['center_y']) < row_threshold:
                current_row.append(current_block)
            else:
                current_row.sort(key=lambda x: x[1]['center_x'])
                row_data = [''] * len(headers)
                has_signature = False
                
                for j, (_, block) in enumerate(current_row):
                    col_idx = min(j, len(headers) - 1)
                    row_data[col_idx] = block['text']
                    if block['text'] in signature_blocks:
                        has_signature = True
                
                if any(cell.strip() for cell in row_data):
                    rows.append(row_data)
                    if has_signature:
                        signature_row_indices.append(len(rows) - 1)
                
                current_row = [current_block]
        
        if current_row:
            current_row.sort(key=lambda x: x[1]['center_x'])
            row_data = [''] * len(headers)
            has_signature = False
            
            for j, (_, block) in enumerate(current_row):
                col_idx = min(j, len(headers) - 1)
                row_data[col_idx] = block['text']
                if block['text'] in signature_blocks:
                    has_signature = True
            
            if any(cell.strip() for cell in row_data):
                rows.append(row_data)
                if has_signature:
                    signature_row_indices.append(len(rows) - 1)
    
    df = pd.DataFrame(rows, columns=headers) if rows else pd.DataFrame(columns=headers)
    return df, signature_row_indices

def create_traditional_table(text_info, signature_blocks):
    text_info.sort(key=lambda x: x['center_y'])

    rows = []
    current_row = [text_info[0]] if text_info else []
    
    if not text_info:
        return pd.DataFrame(), []
    
    heights = [item['height'] for item in text_info]
    avg_height = sum(heights) / len(heights) if heights else 20
    row_height_threshold = avg_height * 0.8
    
    for i in range(1, len(text_info)):
        current_block = text_info[i]
        reference_block = current_row[0]
        
        if abs(current_block['center_y'] - reference_block['center_y']) < row_height_threshold:
            current_row.append(current_block)
        else:
            current_row.sort(key=lambda x: x['center_x'])
            rows.append(current_row)
            current_row = [current_block]

    if current_row:
        current_row.sort(key=lambda x: x['center_x'])
        rows.append(current_row)

    max_cols = max(len(row) for row in rows) if rows else 5
    
    if rows and len(rows[0]) >= 2:
        headers = [block['text'] for block in rows[0]]
        data_rows = rows[1:]
    else:
        headers = [f"Column {i+1}" for i in range(max_cols)]
        data_rows = rows

    while len(headers) < max_cols:
        headers.append(f"Column {len(headers)+1}")
    headers = headers[:max_cols]
    
    table_matrix = []
    signature_row_indices = []
    
    for row_idx, row in enumerate(data_rows):
        row_data = [''] * len(headers)
        row_has_signature = False
        
        for i, block in enumerate(row):
            if i < len(headers):
                row_data[i] = block['text']
                if block['text'] in signature_blocks:
                    row_has_signature = True
        
        if any(cell.strip() for cell in row_data):
            table_matrix.append(row_data)
            if row_has_signature:
                signature_row_indices.append(len(table_matrix) - 1)

    df = pd.DataFrame(table_matrix, columns=headers) if table_matrix else pd.DataFrame(columns=headers)
    return df, signature_row_indices

def save_table_outputs(df, signature_row_indices, all_signatures, text_info, signature_blocks, output_dir):
    csv_path = os.path.join(output_dir, "extracted_table.csv")
    df.to_csv(csv_path, index=False)
    print(f"Table saved to CSV: {csv_path}")

    excel_path = os.path.join(output_dir, "extracted_table.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Table"
    
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    for sig_idx in signature_row_indices:
        excel_row = sig_idx + 2
        for col in range(1, len(df.columns) + 1):
            ws.cell(row=excel_row, column=col).fill = green_fill
    
    wb.save(excel_path)
    print(f"Table saved to Excel: {excel_path}")

    # Save JSON
    json_path = os.path.join(output_dir, "extracted_table.json")
    json_data = {
        "table_data": {
            "headers": df.columns.tolist(),
            "rows": df.values.tolist(),
            "signature_rows": signature_row_indices
        },
        "signature_analysis": {
            "detected_signatures": [{
                "text": sig['block']['text'],
                "confidence": sig.get('confidence', sig.get('signature_score', 0)),
                "type": sig['type']
            } for sig in all_signatures]
        }
    }
    
    with open(json_path, 'w', encoding='utf-8') as json_file:
        json.dump(json_data, json_file, indent=2, ensure_ascii=False)
    
    print(f"Table saved to JSON: {json_path}")
    print(f"\nExtracted {len(df)} rows with {len(df.columns)} columns")

def main():
    image_path = "/home/talgotram/Repos/ioclOCR/input/images/page_2.jpg"
    output_dir = "/home/talgotram/Repos/ioclOCR/output/table_output"

    hf_api_token = os.getenv('HF_API_TOKEN')
    
    if not hf_api_token:
        print("Warning: No HuggingFace API token provided. Using traditional extraction only.")
        print("Get a token from: https://huggingface.co/settings/tokens")
    
    try:
        result = enhanced_table_extraction(
            image_path, 
            output_dir, 
            use_ai=bool(hf_api_token),
            hf_api_token=hf_api_token
        )
        
        if result:
            df, analysis_data = result
            print(f"Successfully extracted table with {len(df)} rows and {len(df.columns)} columns")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()